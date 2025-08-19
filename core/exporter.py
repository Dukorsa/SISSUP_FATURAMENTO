import pandas as pd
import os
from datetime import datetime

# Garante que as bibliotecas de exportação estão disponíveis
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.units import inch
    LIBS_AVAILABLE = True
except ImportError:
    LIBS_AVAILABLE = False

def export_simple_excel(df: pd.DataFrame, path: str, sheet_name: str = 'Relatório'):
    """
    Exporta um DataFrame para um arquivo Excel simples com formatação de cabeçalho.
    """
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6a2e4d", end_color="6a2e4d", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

def export_to_excel(df: pd.DataFrame, path: str, sheet_name: str, totals: dict):
    """
    Exporta um DataFrame para Excel, adicionando uma seção de resumo no final.
    """
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6a2e4d", end_color="6a2e4d", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (max_length + 2)
        if totals:
            totals_row_idx = ws.max_row + 2
            ws.cell(row=totals_row_idx, column=1, value="RESUMO GERAL").font = Font(bold=True)
            for i, (key, val) in enumerate(totals.items(), 1):
                ws.cell(row=totals_row_idx + i, column=1, value=f"{key}:").font = Font(bold=True)
                ws.cell(row=totals_row_idx + i, column=2, value=val)

def _pdf_header_footer(canvas, doc, logo_path, title):
    """
    Cria um cabeçalho e rodapé modernos para o relatório em PDF.
    """
    canvas.saveState()
    page_width, page_height = doc.pagesize

    # --- Cabeçalho Moderno ---
    header_color = colors.HexColor("#6a2e4d")
    canvas.setFillColor(header_color)
    canvas.rect(0, page_height - 0.9*inch, page_width, 0.9*inch, stroke=0, fill=1)

    # Título do Relatório
    canvas.setFillColor(colors.white)
    canvas.setFont('Helvetica-Bold', 18)
    canvas.drawCentredString(page_width / 2.0, page_height - 0.55*inch, title)

    # --- Rodapé ---
    canvas.setFillColor(colors.black)
    canvas.setFont('Helvetica', 9)
    canvas.drawRightString(page_width - doc.rightMargin, 0.5 * inch, f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    canvas.drawCentredString(page_width / 2.0, 0.5 * inch, f"Página {doc.page}")

    canvas.restoreState()

def export_to_pdf(df: pd.DataFrame, path: str, logo_path: str, title: str, totals: dict, pagesize=letter, col_widths=None):
    """
    Exporta o DataFrame para um PDF com layout modernizado e alinhado à esquerda.
    """
    doc = SimpleDocTemplate(path, pagesize=pagesize, topMargin=1.2*inch, bottomMargin=0.8*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # --- Seção de Logo e Resumo ---
    if totals:
        summary_style = ParagraphStyle(name='Summary', fontSize=10, leading=14)
        summary_items = [Paragraph(f"<b>{key}:</b> {value}", summary_style) for key, value in totals.items()]

        max_rows_per_col = 5
        num_cols = (len(summary_items) + max_rows_per_col - 1) // max_rows_per_col
        
        table_data = [[''] * num_cols for _ in range(max_rows_per_col)]
        for i, item in enumerate(summary_items):
            table_data[i % max_rows_per_col][i // max_rows_per_col] = item

        col_widths_summary = [2.8 * inch] * num_cols
        summary_table = Table(table_data, colWidths=col_widths_summary)
        summary_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0)]))
        
        logo_image = Image(logo_path, width=1.8*inch, height=1.8*inch, kind='bound', hAlign='CENTER') if os.path.exists(logo_path) else Spacer(0, 0)

        # CORREÇÃO: Adicionado hAlign='LEFT' para forçar o alinhamento
        layout_table = Table([[logo_image, summary_table]], colWidths=[2.2*inch, sum(col_widths_summary) + 0.1*inch], hAlign='LEFT')
        layout_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
        elements.append(layout_table)

    elements.append(Spacer(1, 0.3*inch))
    
    # --- Tabela Principal de Dados ---
    header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    body_style_center = ParagraphStyle(name='BodyStyleCenter', fontSize=8, alignment=TA_CENTER, leading=10)
    body_style_left = ParagraphStyle(name='BodyStyleLeft', fontSize=8, alignment=TA_LEFT, leading=10)
    styled_data = [[Paragraph(col, header_style) for col in df.columns]]

    center_cols = ['Nº APAC', 'HD', 'Extras', 'CDL', 'Observação', 'Número da Guia', 'Matrícula', 'Lote', 'Quant.', 'Total', 'Data Início', 'Data Final']
    for _, row in df.iterrows():
        row_data = [Paragraph(str(item), body_style_center if col_name in center_cols else body_style_left) for col_name, item in row.items()]
        styled_data.append(row_data)

    table = Table(styled_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6a2e4d")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(table)
    
    header_footer_with_args = lambda canvas, doc: _pdf_header_footer(canvas, doc, logo_path, title)
    doc.build(elements, onFirstPage=header_footer_with_args, onLaterPages=header_footer_with_args)

def export_fistulas_to_excel(df: pd.DataFrame, path: str, totals: dict):
    """
    Exporta o relatório de fístulas para Excel com uma seção de resumo.
    """
    sheet_name = "Procedimentos FAV"
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6a2e4d", end_color="6a2e4d", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (max_length + 2)
        if totals:
            totals_row_idx = ws.max_row + 2
            ws.cell(row=totals_row_idx, column=1, value="RESUMO DE PROCEDIMENTOS").font = Font(bold=True)
            for i, (key, val) in enumerate(totals.items(), 1):
                ws.cell(row=totals_row_idx + i, column=1, value=f"{key}:").font = Font(bold=True)
                ws.cell(row=totals_row_idx + i, column=2, value=val)

def export_fistulas_to_pdf(df: pd.DataFrame, path: str, logo_path: str, totals: dict):
    """
    Exporta o relatório de fístulas para PDF com layout modernizado e alinhado à esquerda.
    """
    doc = SimpleDocTemplate(path, pagesize=letter, topMargin=1.2*inch, bottomMargin=0.8*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    title = "Relatório de Procedimentos FAV"

    # --- Seção de Logo e Resumo ---
    if totals:
        summary_style = ParagraphStyle(name='Summary', fontSize=10, leading=14)
        summary_items = [Paragraph(f"<b>{key}:</b> {value}", summary_style) for key, value in totals.items()]
        
        max_rows_per_col = 5
        num_cols = (len(summary_items) + max_rows_per_col - 1) // max_rows_per_col
        
        table_data = [[''] * num_cols for _ in range(max_rows_per_col)]
        for i, item in enumerate(summary_items):
            table_data[i % max_rows_per_col][i // max_rows_per_col] = item

        col_widths_summary = [3.0 * inch] * num_cols
        summary_table = Table(table_data, colWidths=col_widths_summary)
        summary_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0)]))
        
        logo_image = Image(logo_path, width=1.8*inch, height=1.8*inch, kind='bound', hAlign='CENTER') if os.path.exists(logo_path) else Spacer(0, 0)
        
        # CORREÇÃO: Adicionado hAlign='LEFT' para forçar o alinhamento
        layout_table = Table([[logo_image, summary_table]], colWidths=[2.2*inch, sum(col_widths_summary) + 0.1*inch], hAlign='LEFT')
        layout_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
        elements.append(layout_table)
    
    elements.append(Spacer(1, 0.3*inch))
    
    # --- Tabela de Dados ---
    header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    body_style_center = ParagraphStyle(name='BodyStyleCenter', fontSize=8, alignment=TA_CENTER, leading=10)
    body_style_left = ParagraphStyle(name='BodyStyleLeft', fontSize=8, alignment=TA_LEFT, leading=10)
    styled_data = [[Paragraph(col, header_style) for col in df.columns]]
    
    for _, row in df.iterrows():
        row_data = [Paragraph(str(item), body_style_left if col_name == 'Nome' else body_style_center) for col_name, item in row.items()]
        styled_data.append(row_data)
        
    table = Table(styled_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6a2e4d")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(table)
    
    header_footer_with_args = lambda canvas, doc: _pdf_header_footer(canvas, doc, logo_path, title)
    doc.build(elements, onFirstPage=header_footer_with_args, onLaterPages=header_footer_with_args)

def export_continuidade_to_pdf(df: pd.DataFrame, path: str, logo_path: str, title: str):
    """
    Exporta o relatório de continuidade para PDF com layout modernizado e alinhado à esquerda.
    """
    doc = SimpleDocTemplate(path, pagesize=letter, topMargin=1.2*inch, bottomMargin=0.8*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # --- Seção de Logo e Resumo ---
    summary_style = ParagraphStyle(name='Summary', fontSize=10, leading=14)
    summary_text = f"<b>Total de Pacientes em Continuidade:</b> {len(df)}"
    summary_paragraph = Paragraph(summary_text, summary_style)

    logo_image = Image(logo_path, width=1.8*inch, height=1.8*inch, kind='bound', hAlign='CENTER') if os.path.exists(logo_path) else Spacer(0, 0)

    # CORREÇÃO: Adicionado hAlign='LEFT' para forçar o alinhamento
    layout_table = Table([[logo_image, summary_paragraph]], colWidths=[2.2*inch, 5.3*inch], hAlign='LEFT')
    layout_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (1, 0), (1, 0), 0)]))
    elements.append(layout_table)
    
    elements.append(Spacer(1, 0.3*inch))
    
    # --- Tabela de Dados ---
    header_style = ParagraphStyle(name='HeaderStyle', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    body_style_center = ParagraphStyle(name='BodyStyleCenter', fontSize=8, alignment=TA_CENTER, leading=10)
    body_style_left = ParagraphStyle(name='BodyStyleLeft', fontSize=8, alignment=TA_LEFT, leading=10)
    styled_data = [[Paragraph(col, header_style) for col in df.columns]]
    
    for _, row in df.iterrows():
        row_data = [Paragraph(str(item), body_style_left if col_name == 'Nome' else body_style_center) for col_name, item in row.items()]
        styled_data.append(row_data)
        
    table = Table(styled_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6a2e4d")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(table)
    
    header_footer_with_args = lambda canvas, doc: _pdf_header_footer(canvas, doc, logo_path, title)
    doc.build(elements, onFirstPage=header_footer_with_args, onLaterPages=header_footer_with_args)